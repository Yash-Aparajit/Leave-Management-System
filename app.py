# app.py
import os
import io
import csv
import shutil
import uuid
import json
from openpyxl import load_workbook
from functools import wraps
from datetime import date, datetime as dt, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import PatternFill
from io import StringIO, BytesIO
import pandas as pd        # type: ignore

from models import (
    db, Employee, LeaveType, LeaveEntry, Transaction, User,
    CompOffRecord, EarlyLateRecord,OutdoorDuty 
)

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    send_file, session
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, and_, or_

# Excel support
from openpyxl import Workbook

# ---------- App setup ----------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "app.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
BACKUPS_FOLDER = os.path.join(BASE_DIR, "backups")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUPS_FOLDER, exist_ok=True)

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "change-this-secret-string-to-something-secret"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["SYSTEM_VERSION"] = "1.0.4"
app.config["BUILD_DATE"] = "2025-11-25"

db.init_app(app)

# ---------- Roles & permissions ----------
ROLE_PERMISSIONS = {
    "admin_1": {
        "can_override": False,
        "can_restore_db": False,
        "can_delete_leave": False,
        "can_edit_employee": True,
        "can_set_manual_balance": False,
    },
    "admin_master": {
        "can_override": True,
        "can_restore_db": True,
        "can_delete_leave": True,
        "can_edit_employee": True,
        "can_set_manual_balance": True,
    },
    "developer": {
        "can_override": True,
        "can_restore_db": True,
        "can_delete_leave": True,
        "can_edit_employee": True,
        "can_set_manual_balance": True,
    },
}

def normalize_role(role):
    if not role:
        return None

    # normalize case and whitespace
    role = role.strip().lower()

    if role == "admin_override":
        return "admin_master"

    return role



def has_permission(permission_name):
    uid = session.get("user_id")
    if not uid:
        return False
    user = User.query.get(uid)
    if not user:
        return False
    role = normalize_role(user.role)
    perms = ROLE_PERMISSIONS.get(role, {})
    return perms.get(permission_name, False)


def require_permission(permission_name):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not has_permission(permission_name):
                flash("Permission denied", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)

        return wrapper

    return decorator

# ---------- session & auth ----------
@app.before_request
def validate_session_token():
    endpoint = (request.endpoint or "")
    public_endpoints = ("login", "create_tables_and_seed", "static")
    if endpoint.startswith("static") or endpoint in public_endpoints:
        return
    uid = session.get("user_id")
    token = session.get("session_token")
    if not uid or not token:
        return
    try:
        user = db.session.get(User, uid)
    except Exception:
        user = User.query.get(uid)
    if not user or not user.session_token or user.session_token != token:
        session.clear()
        flash("Session expired — please log in again.", "info")
        return redirect(url_for("login"))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)

    return decorated

@app.template_filter("nice_date")
def nice_date(value):
    if not value:
        return ""
    try:
        # works for both date and datetime
        return value.strftime("%d/%m/%Y")
    except Exception:
        return str(value)

@app.context_processor
def inject_permissions():
    return dict(has_permission=has_permission, session=session)

def is_employee_left(emp, ref_date=None):
    """
    Canonical check: returns True if employee is considered left
    as of ref_date (or today if not provided).
    """
    if not emp:
        return False

    ref_date = ref_date or date.today()

    # 1. Explicit status
    status = getattr(emp, "status", None)
    if status and str(status).lower() == "left":
        return True

    # 2. left_date check
    left_date = getattr(emp, "left_date", None)
    if left_date:
        try:
            if isinstance(left_date, str):
                left_date = dt.strptime(left_date.strip(), "%Y-%m-%d").date()
            elif hasattr(left_date, "date"):
                left_date = left_date.date()
            if isinstance(left_date, date) and left_date <= ref_date:
                return True
        except Exception:
            pass

    # 3. Optional legacy flag
    is_active = getattr(emp, "is_active", None)
    if is_active is not None and not bool(is_active):
        return True

    return False

# ---------- accrual & balance helpers ----------
def apply_missing_accruals_for_employee(emp):
    """
    Ensure monthly ACCRUAL transactions exist from hire_date (month-start)
    up to the last completed month.

    Behavior changes:
    - Skip creating accruals for months that are entirely after an employee's left_date.
    - If emp.status == 'left' but left_date is missing, we treat them as left as of today (no further accruals).
    - Optional pro-rata logic for month when left_date falls inside the month is included but commented (enable if you want pro-rata).
    """
    if not getattr(emp, "hire_date", None):
        return

    today = date.today()
    cur = emp.hire_date.replace(day=1)
    last_month = today.replace(day=1)

    # helper: safe float conversion
    def safe_float(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    # Normalize left_date if present (may be None, string, date, or datetime)
    raw_left = getattr(emp, "left_date", None)
    left_date = None
    if raw_left:
        try:
            if isinstance(raw_left, str):
                # expect YYYY-MM-DD (init_db stores DATE) — fallback tolerant parse
                left_date = dt.strptime(raw_left.strip(), "%Y-%m-%d").date()
            elif hasattr(raw_left, "date"):  
                left_date = raw_left.date()
            else:
                left_date = raw_left
        except Exception:
            left_date = None

    # If status explicitly 'left' and no left_date, treat as left as of today (skip future accruals)
    status_val = getattr(emp, "status", None)
    if (status_val is not None) and (str(status_val).lower() == "left") and (left_date is None):
        left_date = date.today()

    # main loop: walk months from hire_date (month-start) up to last_month
    while cur <= last_month:
        # month_end is last day of the month 'cur'
        if cur.month == 12:
            month_end = cur.replace(year=cur.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = cur.replace(month=cur.month + 1, day=1) - timedelta(days=1)

        #  - if left_date <= (month_start - 1) -> employee already left BEFORE this month -> skip / stop creating accruals
        if left_date:
            # employee left before or on the month end -> do not create a *full* accrual for this month
            if left_date <= month_end:
                # If left_date is before the month_start, we should stop completely (no accruals for this month or after)
                if left_date < cur:
                    # left before this month started -> stop loop entirely
                    break
                else:
                    break

        # Normal path: create full accrual if record not present
        period = cur.strftime("%Y-%m")
        exists = Transaction.query.filter_by(employee_id=emp.id, type="ACCRUAL", period=period).first()
        if not exists:
            rate = safe_float(getattr(emp, "accrual_rate", 0.0))
            amount = round(rate, 2)

            created_by = None
            try:
                created_by = session.get("user_id")
            except Exception:
                created_by = None

            t = Transaction(
                employee_id=emp.id,
                type="ACCRUAL",
                period=period,
                amount=amount,
                note=f"Auto-accrual for {period}",
                created_by=created_by,
            )
            db.session.add(t)

        # increment month (move to first day of next month)
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise


def recalc_accruals_for_promotion(emp, old_rate, new_rate, promotion_date):
    """
    Rebuild all ACCRUAL transactions for an employee given a promotion.

    Rules:
    - Months BEFORE promotion use old_rate.
    - Months AFTER promotion use new_rate.
    - If promotion_date.day == 1 -> that month uses new_rate.
    - If promotion_date.day > 1 -> next month uses new_rate.
    """
    if not emp.hire_date:
        return

    switchover = None
    if promotion_date:
        if promotion_date.day == 1:
            switchover = date(promotion_date.year, promotion_date.month, 1)
        else:
            # Next month
            if promotion_date.month == 12:
                switchover = date(promotion_date.year + 1, 1, 1)
            else:
                switchover = date(promotion_date.year, promotion_date.month + 1, 1)

    # Delete all previous accruals for this employee
    Transaction.query.filter_by(employee_id=emp.id, type="ACCRUAL").delete()
    db.session.flush()

    today = date.today()
    cur = emp.hire_date.replace(day=1)
    last_month = today.replace(day=1)

    while cur <= last_month:
        if switchover:
            rate = new_rate if cur >= switchover else old_rate
        else:
            rate = new_rate

        period = cur.strftime("%Y-%m")
        t = Transaction(
            employee_id=emp.id,
            type="ACCRUAL",
            period=period,
            amount=round(float(rate or 0), 2),
            note=f"Recalc accrual for {period} (promo logic)",
            created_by=session.get("user_id"),
        )
        db.session.add(t)

        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)

    db.session.commit()

def compute_balance(emp):
    """
    Effective leave balance = sum of all transactions for this employee.
    Manual overrides are applied as MANUAL_OVERRIDE transactions,
    so we ALWAYS compute from the ledger, not from emp.manual_balance.
    """
    try:
        apply_missing_accruals_for_employee(emp)
    except Exception:
        pass
    total = (
        db.session.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.employee_id == emp.id)
        .scalar()
    )
    return round(float(total or 0), 2)


# ---------- Authentication routes ----------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            role = normalize_role(user.role)
            ALLOWED_ROLES = ("admin_1", "admin_master", "developer")
            if user.role != role:
                user.role = role
            token = uuid.uuid4().hex
            user.session_token = token
            db.session.commit()
            session["session_token"] = token
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = role
            if getattr(user, "force_password_change", False):
                flash(
                    "You must change the default password before continuing.", "info"
                )
                return redirect(url_for("change_password"))
            flash("Logged in", "success")
            return redirect(url_for("index"))
        flash("Invalid credentials", "danger")
    return render_template("login.html")



@app.route("/logout")
def logout():
    try:
        uid = session.get("user_id")
        if uid:
            user = User.query.get(uid)
            if user:
                user.session_token = None
                db.session.commit()
    except Exception:
        pass
    session.clear()
    flash("Logged out", "info")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = User.query.get(session.get("user_id"))
    if not user:
        flash("User not found", "danger")
        return redirect(url_for("login"))
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not check_password_hash(user.password_hash, current):
            flash("Current password is incorrect", "danger")
            return redirect(url_for("change_password"))
        if len(new) < 8:
            flash("New password must be at least 8 characters", "danger")
            return redirect(url_for("change_password"))
        if new != confirm:
            flash("New password and confirmation do not match", "danger")
            return redirect(url_for("change_password"))
        user.password_hash = generate_password_hash(new)
        user.force_password_change = False
        user.session_token = None
        db.session.commit()
        session.clear()
        flash("Password changed. Please log in again.", "success")
        return redirect(url_for("login"))
    return render_template("change_password.html")


# ---------- Dashboard & general pages ----------
@app.route('/')
@login_required
def index():
    """
    Dashboard / home page.
    Defensive: catches internal errors and always returns a valid response.
    """
    try:
        today = date.today()

        total_employees = Employee.query.filter(Employee.status == 'active').count()

        week_start = today - timedelta(days=today.weekday())
        week_count = LeaveEntry.query.filter(LeaveEntry.date_from >= week_start).count()

        month_count = LeaveEntry.query.filter(func.strftime('%Y-%m', LeaveEntry.date_from) == today.strftime('%Y-%m')).count()
        year_count = LeaveEntry.query.filter(func.strftime('%Y', LeaveEntry.date_from) == str(today.year)).count()

        leave_types = LeaveType.query.all()
        counts_by_type = {}
        for lt in leave_types:
            try:
                counts_by_type[lt.name] = LeaveEntry.query.filter(LeaveEntry.leave_type_id == lt.id).count()
            except Exception:
                counts_by_type[lt.name] = 0

        employees = Employee.query.filter(Employee.status == 'active').order_by(Employee.last_name).limit(50).all()

        balances = {}
        for e in employees:
            try:
                balances[e.id] = compute_balance(e)
            except Exception:
                balances[e.id] = 0

        # finally render index template
        return render_template(
            'index.html',
            total_employees=total_employees,
            week_count=week_count,
            month_count=month_count,
            year_count=year_count,
            employees=employees,
            balances=balances,
            counts_by_type=counts_by_type,
            leave_types=leave_types,
            system_version=app.config.get('SYSTEM_VERSION', ''),
            build_date=app.config.get('BUILD_DATE', '')
        )

    except Exception as exc:
        import traceback
        traceback.print_exc()
        flash(f'Error loading dashboard: {exc}', 'danger')

        try:
            return render_template('index.html',
                                   total_employees=0,
                                   week_count=0,
                                   month_count=0,
                                   year_count=0,
                                   employees=[],
                                   balances={},
                                   counts_by_type={},
                                   leave_types=[],
                                   system_version=app.config.get('SYSTEM_VERSION', ''),
                                   build_date=app.config.get('BUILD_DATE', '')
                                   )
        except Exception:
            return "Dashboard error. Check server logs."

@app.route("/help")
@login_required
def help_page():
    return render_template(
        "help.html",
        system_version=app.config.get("SYSTEM_VERSION", ""),
        build_date=app.config.get("BUILD_DATE", ""),
    )


    
# ---------- Employee listing & search ----------
@app.route("/employees")
@login_required
def employees():
    return redirect(url_for("employees_list"))


@app.route("/employees_list")
@login_required
def employees_list():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "active")
    try:
        page = int(request.args.get("page", "1"))
    except Exception:
        page = 1
    per_page = 40

    base = Employee.query
    if status:
        base = base.filter(Employee.status == status)
    if q:
        base = base.filter(
            or_(
                Employee.employee_id.ilike(f"%{q}%"),
                Employee.first_name.ilike(f"%{q}%"),
                Employee.middle_name.ilike(f"%{q}%"),
                Employee.last_name.ilike(f"%{q}%"),
            )
        )
    total = base.count()
    employees = (
        base.order_by(Employee.employee_id)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    balances = {}
    try:
        for e in employees:
            try:
                balances[e.id] = compute_balance(e)
            except Exception:
                balances[e.id] = 0
    except Exception:
        balances = {e.id: 0 for e in employees}

    return render_template(
        "employees_list.html",
        employees=employees,
        q=q,
        status=status,
        page=page,
        per_page=per_page,
        total=total,
        balances=balances,
    )


@app.route("/search")
@login_required
def search():
    q = request.args.get("q", "").strip()
    if not q:
        return redirect(url_for("index"))
    emp = Employee.query.filter_by(employee_id=q).first()
    if emp:
        return redirect(url_for("employee_detail", emp_id=emp.id))
    parts = q.split()
    if len(parts) == 1:
        employees = Employee.query.filter(
            or_(
                Employee.first_name.ilike(f"%{q}%"),
                Employee.middle_name.ilike(f"%{q}%"),
                Employee.last_name.ilike(f"%{q}%"),
            )
        ).all()
    else:
        employees = Employee.query.filter(
            and_(
                Employee.first_name.ilike(f"%{parts[0]}%"),
                Employee.last_name.ilike(f"%{parts[-1]}%"),
            )
        ).all()
    return render_template("employees.html", employees=employees)


# ---------- Add / Remove / Promote / Edit employee ----------
@app.route("/employees/add", methods=["GET", "POST"])
@login_required
def add_employee():
    departments = [
        "Purchase",
        "ME",
        "HR Admin",
        "Production",
        "Quality",
        "Store",
        "Other",
    ]
    if request.method == "POST":
        emp_id = request.form.get("employee_id", "").strip()
        first = request.form.get("first_name", "").strip()
        middle = request.form.get("middle_name", "").strip() or None
        last = request.form.get("last_name", "").strip()
        try:
            hire_date = dt.strptime(request.form.get("hire_date"), "%Y-%m-%d").date()
        except Exception:
            flash("Invalid hire date", "danger")
            return redirect(url_for("add_employee"))
        accrual_rate = float(request.form.get("accrual_rate", 1.5))
        plant_location = request.form.get('plant_location', '').strip() or None
        department = request.form.get("department", "").strip() or None
        designation = request.form.get("designation", "").strip() or None
        contact = request.form.get("contact_number", "").strip() or None
        emergency = request.form.get("emergency_number", "").strip() or None

        def is_valid_phone(n):
            return n.isdigit() and len(n) == 10

        if not contact or not is_valid_phone(contact):
            flash("Contact number must be exactly 10 digits.", "danger")
            return redirect(url_for("add_employee"))

        if Employee.query.filter_by(contact_number=contact).first():
            flash("This contact number already exists for another employee.", "danger")
            return redirect(url_for("add_employee"))

        if emergency and not is_valid_phone(emergency):
            flash("Emergency number must be exactly 10 digits.", "danger")
            return redirect(url_for("add_employee"))

        if Employee.query.filter_by(employee_id=emp_id).first():
            flash("Employee ID already exists", "danger")
            return redirect(url_for("add_employee"))
        emp = Employee(
            employee_id=emp_id,
            first_name=first,
            middle_name=middle,
            last_name=last,
            hire_date=hire_date,
            accrual_rate=accrual_rate,
            status="active",
            plant_location=plant_location,
            department=department,
            designation=designation,
            contact_number=contact,
            emergency_number=emergency,
        )
        db.session.add(emp)
        db.session.commit()
        flash("Employee added", "success")
        return redirect(url_for("employees_list"))
    return render_template(
        "add_employee.html",
        departments=["Purchase", "ME", "HR Admin", "Production", "Quality", "Store", "Other"],
    )


@app.route("/employees/remove", methods=["GET", "POST"])
@login_required
def remove_employee():
    if request.method == "POST":
        emp_id = request.form.get("employee_id", "").strip()
        emp = Employee.query.filter_by(employee_id=emp_id).first()
        if not emp:
            flash("Employee not found", "danger")
            return redirect(url_for("remove_employee"))
        emp.status = "left"
        emp.left_date = date.today()
        db.session.commit()
        flash(f"{emp.employee_id} marked as left", "success")
        return redirect(url_for("employees_list"))
    return render_template("remove_employee.html")

from datetime import date, datetime as dt  

@app.route('/promote', methods=['GET','POST'])
@login_required
def promote_employee():
    """
    Promote an employee: set new accrual_rate and promotion_date.
    """
    if request.method == 'POST':
        emp_id = request.form.get('employee_id','').strip()
        try:
            new_rate = float(request.form.get('new_rate'))
        except Exception:
            flash('Invalid new rate', 'danger')
            return redirect(url_for('promote_employee'))
        try:
            promo_date = dt.strptime(request.form.get('effective_date'), '%Y-%m-%d').date()
        except Exception:
            flash('Invalid effective date', 'danger')
            return redirect(url_for('promote_employee'))

        emp = Employee.query.filter_by(employee_id=emp_id).first()
        if not emp:
            flash('Employee not found', 'danger')
            return redirect(url_for('promote_employee'))

        if is_employee_left(emp, promo_date):
            flash(f'Employee {emp.employee_id} has left — promotion not allowed.', 'danger')
            return redirect(url_for('promote_employee'))

        # Save old rate for auditing
        old_rate = float(emp.accrual_rate or 0.0)

        # Compute promotion effective month-start:
        if promo_date.day == 1:
            effective_month_start = date(promo_date.year, promo_date.month, 1)
        else:
            if promo_date.month == 12:
                effective_month_start = date(promo_date.year + 1, 1, 1)
            else:
                effective_month_start = date(promo_date.year, promo_date.month + 1, 1)

        # Update the employee record with new values
        emp.promotion_date = promo_date
        emp.accrual_rate = new_rate
        db.session.commit()

        prom_note = f'Promotion recorded by {session.get("username")}: {old_rate} -> {new_rate}, effective {effective_month_start.isoformat()}'
        tr_prom = Transaction(
            employee_id=emp.id,
            type='PROMOTION',
            period=None,
            amount=0.0,
            note=prom_note,
            created_by=session.get('user_id')
        )
        db.session.add(tr_prom)
        db.session.commit()

        try:
            if 'recalc_accruals_for_promotion' in globals() and callable(recalc_accruals_for_promotion):
                recalc_accruals_for_promotion(emp, old_rate, new_rate, promo_date)
        except Exception:
            flash('Warning: recalc_accruals_for_promotion helper failed (non-fatal).', 'warning')

        today = date.today()
        last_month = today.replace(day=1)
        cur = effective_month_start
        total_delta = 0.0

        while cur <= last_month:
            period = cur.strftime('%Y-%m')
            actual = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
                Transaction.employee_id == emp.id,
                Transaction.type == 'ACCRUAL',
                Transaction.period == period
            ).scalar() or 0.0
            expected = round(float(new_rate), 2)
            delta = round(expected - float(actual), 2)
            total_delta += delta

            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1, day=1)
            else:
                cur = cur.replace(month=cur.month + 1, day=1)

        if abs(total_delta) > 0.0001:
            adj_note = f'Promotion adjustment for {emp.employee_id}: {old_rate}->{new_rate} effective {effective_month_start.isoformat()}'
            tr_adj = Transaction(
                employee_id=emp.id,
                type='PROMOTION_ADJUST',
                period=None,
                amount=round(total_delta, 2),
                note=adj_note,
                created_by=session.get('user_id')
            )
            db.session.add(tr_adj)
            db.session.commit()

        flash(f'Promotion recorded for {emp.employee_id} (effective {effective_month_start.isoformat()})', 'success')
        return redirect(url_for('employees_list'))

    return render_template('promote.html')

@app.route("/employees/edit/<int:emp_id>", methods=["GET", "POST"])
@login_required
def edit_employee(emp_id):
    """
    Edit employee profile: allows editing of contact, emergency contact,
    department, designation, accrual rate, hire/promotion dates, status,
    and (for admin_master/developer) manual_balance-based one-time override.
    """
    emp = Employee.query.get_or_404(emp_id)

    if not has_permission("can_edit_employee") and session.get("role") not in (
        "developer",
        "admin_master",
    ):
        flash("No permission to edit employee", "danger")
        return redirect(url_for("employee_detail", emp_id=emp.id))

    departments = [
        "Purchase",
        "ME",
        "HR Admin",
        "Production",
        "Quality",
        "Store",
        "Other",
    ]

    if request.method == "POST":
        # Basic profile updates
        emp.employee_id = request.form.get("employee_id", emp.employee_id).strip()
        emp.first_name = request.form.get("first_name", emp.first_name).strip()
        emp.middle_name = request.form.get("middle_name", "").strip() or None
        emp.last_name = request.form.get("last_name", emp.last_name).strip()

        try:
            hire_date_str = request.form.get("hire_date")
            if hire_date_str:
                emp.hire_date = dt.strptime(hire_date_str, "%Y-%m-%d").date()
        except Exception:
            flash("Invalid hire date", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        try:
            promo_str = request.form.get("promotion_date")
            emp.promotion_date = (
                dt.strptime(promo_str, "%Y-%m-%d").date() if promo_str else None
            )
        except Exception:
            flash("Invalid promotion date", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        try:
            emp.accrual_rate = float(
                request.form.get("accrual_rate", emp.accrual_rate)
            )
        except Exception:
            flash("Invalid accrual rate", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        emp.status = request.form.get("status", emp.status)

        emp.plant_location = request.form.get("plant_location", "").strip() or None
        emp.department = request.form.get("department", "").strip() or None
        emp.designation = request.form.get("designation", "").strip() or None
        emp.contact_number = request.form.get("contact_number", "").strip() or None
        emp.emergency_number = (
            request.form.get("emergency_number", "").strip() or None
        )

        def is_valid_phone(n):
            return n.isdigit() and len(n) == 10

        contact = emp.contact_number
        emergency = emp.emergency_number

        if not contact or not is_valid_phone(contact):
            flash("Contact number must be exactly 10 digits.", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        existing = Employee.query.filter(
            Employee.contact_number == contact,
            Employee.id != emp.id
        ).first()

        if existing:
            flash("This contact number already exists for another employee.", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        if emergency and not is_valid_phone(emergency):
            flash("Emergency number must be exactly 10 digits.", "danger")
            return redirect(url_for("edit_employee", emp_id=emp.id))

        # ----- Manual balance override as a one-time correction -----
        new_manual_raw = request.form.get("manual_balance", "").strip()
        manual_override_requested = False
        new_manual_val = None
        previous_balance = None

    # 🔒 Block manual balance override for LEFT employees (developer only)
        if new_manual_raw != "" and is_employee_left(emp) and session.get("role") != "developer":
            flash(
                "Manual balance correction for a left employee is restricted to developer only.",
                "danger"
            )
            return redirect(url_for("edit_employee", emp_id=emp.id))


        if new_manual_raw != "":
            # Only developer and admin_master can use manual override
            if session.get("role") not in ("developer", "admin_master"):
                flash("Permission denied for manual balance change", "danger")
                return redirect(url_for("edit_employee", emp_id=emp.id))
            try:
                new_manual_val = float(new_manual_raw)
            except Exception:
                flash("Invalid manual balance value", "danger")
                return redirect(url_for("edit_employee", emp_id=emp.id))

            try:
                previous_balance = compute_balance(emp)
            except Exception:
                previous_balance = 0.0

            manual_override_requested = True

            emp.manual_balance = new_manual_val

        db.session.commit()

        if manual_override_requested and previous_balance is not None:
            try:
                delta = round(float(new_manual_val) - float(previous_balance), 2)
                if abs(delta) > 0.0001:
                    note = (
                        f"MANUAL_OVERRIDE by {session.get('username')}: "
                        f"previous={previous_balance}, new={new_manual_val}"
                    )
                    tr = Transaction(
                        employee_id=emp.id,
                        type="MANUAL_OVERRIDE",
                        period=None,
                        amount=delta,
                        reference_id=None,
                        note=note,
                        created_by=session.get("user_id"),
                    )
                    db.session.add(tr)
                    db.session.commit()
                else:
                    pass
            except Exception as e:
                db.session.rollback()
                flash(
                    "Warning: manual override saved but audit record failed: " + str(e),
                    "warning",
                )

        flash("Employee updated", "success")
        return redirect(url_for("employee_detail", emp_id=emp.id))

    return render_template("edit_employee.html", e=emp, departments=departments)



# ---------- Record / Edit / Delete leave ----------
@app.route('/leave/record', methods=['GET','POST'])
@login_required
def record_leave():
    leave_types = LeaveType.query.all()
    if request.method == 'POST':
        emp_code = request.form.get('employee_code','').strip()
        emp = Employee.query.filter_by(employee_id=emp_code).first()
        if not emp:
            flash('Employee not found', 'danger')
            return redirect(url_for('record_leave'))

        # Block if employee left/inactive
        if emp.status != 'active':
            flash('Cannot record leave for an employee who has left / is inactive.', 'danger')
            return redirect(url_for('record_leave'))

        try:
            lt_id = int(request.form.get('leave_type_id'))
            date_from = dt.strptime(request.form.get('date_from'), '%Y-%m-%d').date()
            date_to = dt.strptime(request.form.get('date_to'), '%Y-%m-%d').date()
            days = float(request.form.get('days'))
        except Exception:
            flash('Invalid leave data', 'danger')
            return redirect(url_for('record_leave'))

        situation = request.form.get('situation','').strip() or None

        # approver (who approved the leave) - optional
        approver = request.form.get('approver','').strip() or None

        # recorder_name: who entered this record into the system (new)
        recorder_name = request.form.get('recorder_name','').strip() or None

        reason = request.form.get('reason','').strip()
        lt = LeaveType.query.get(lt_id)

        six_months_after = emp.hire_date + relativedelta(months=6) if emp.hire_date else date.today()

        # If paid leave within 6 months -> require override permission
        if lt and getattr(lt, 'is_paid', True) and date_from < six_months_after:
            if has_permission('can_override') or session.get('role') in ('developer','admin_master'):
                le = LeaveEntry(
                    employee_id=emp.id,
                    date_from=date_from,
                    date_to=date_to,
                    days=days,
                    leave_type_id=lt_id,
                    situation=situation,
                    reason=reason,
                    approver=approver,
                    created_by=session.get('user_id')
                )

                try:
                    setattr(le, 'approver', approver or None)
                except Exception:
                    pass
                try:
                    setattr(le, 'recorder_name', recorder_name or None)
                except Exception:
                    pass

                db.session.add(le)
                db.session.flush()

                note_txt = f'Leave {date_from} to {date_to}.'
                if approver:
                    note_txt += f' Approver: {approver}.'
                if situation:
                    note_txt += f' Situation: {situation}.'
                if recorder_name:
                    note_txt += f' Recorded by: {recorder_name}.'

                tr = Transaction(
                    employee_id=emp.id,
                    type='LEAVE_TAKEN',
                    amount=round(-abs(days), 2),
                    reference_id=le.id,
                    note=note_txt + ' (OVERRIDDEN)',
                    created_by=session.get('user_id')
                )
                db.session.add(tr)
                ov = Transaction(
                    employee_id=emp.id,
                    type='OVERRIDE',
                    period=None,
                    amount=0.0,
                    reference_id=le.id,
                    note=(f'Override for paid leave within 6 months by user {session.get("username")}'
                          + (f' — Approver: {approver}' if approver else '')
                          + (f' — Situation: {situation}' if situation else '')
                          + (f' — Recorder: {recorder_name}' if recorder_name else '')),
                    created_by=session.get('user_id')
                )
                db.session.add(ov)
                db.session.commit()
                flash('Leave recorded with override', 'success')
                return redirect(url_for('employee_detail', emp_id=emp.id))
            else:
                flash(f'Paid leave not allowed until {six_months_after}. Ask Admin with override access.', 'danger')
                return redirect(url_for('record_leave'))

        le = LeaveEntry(
            employee_id=emp.id,
            date_from=date_from,
            date_to=date_to,
            days=days,
            leave_type_id=lt_id,
            situation=situation,
            reason=reason,
            created_by=session.get('user_id')
        )
        try:
            setattr(le, 'approver', approver or None)
        except Exception:
            pass
        try:
            setattr(le, 'recorder_name', recorder_name or None)
        except Exception:
            pass

        db.session.add(le)
        db.session.flush()

        # Paid vs Unpaid determination via leave type boolean is_paid
        is_paid = bool(lt and getattr(lt, 'is_paid', True))
        amount = round(-abs(days), 2) if is_paid else 0.0

        note_txt = f'Leave {date_from} to {date_to}.'
        if approver:
            note_txt += f' Approver: {approver}.'
        if situation:
            note_txt += f' Situation: {situation}.'
        if recorder_name:
            note_txt += f' Recorded by: {recorder_name}.'
        if not is_paid:
            note_txt += ' (UNPAID – no balance deduction)'

        tr = Transaction(
            employee_id=emp.id,
            type='LEAVE_TAKEN',
            amount=amount,
            reference_id=le.id,
            note=note_txt,
            created_by=session.get('user_id')
        )
        db.session.add(tr)
        db.session.commit()
        flash('Leave recorded', 'success')
        return redirect(url_for('employee_detail', emp_id=emp.id))

    return render_template('record_leave_v2.html', leave_types=leave_types)

@app.route('/leave/edit/<int:leave_id>', methods=['GET','POST'])
@login_required
def edit_leave(leave_id):
    le = LeaveEntry.query.get_or_404(leave_id)
    emp = Employee.query.get_or_404(le.employee_id)
    if session.get("role") != "developer":
        flash("Your are not allowed to edit leave records.", "danger")
        return redirect(url_for("employee_detail", emp_id=emp.id))

    if request.method == 'POST':
        try:
            date_from = dt.strptime(request.form.get('date_from'), '%Y-%m-%d').date()
            date_to = dt.strptime(request.form.get('date_to'), '%Y-%m-%d').date()
            days = float(request.form.get('days'))
            lt_id = int(request.form.get('leave_type_id'))
            reason = request.form.get('reason','').strip()
            approver = request.form.get('approver','').strip()
            recorder_name = request.form.get('recorder_name','').strip()   # NEW
        except Exception:
            flash('Invalid data', 'danger')
            return redirect(url_for('edit_leave', leave_id=leave_id))
        le.date_from = date_from
        le.date_to = date_to
        le.days = days
        le.leave_type_id = lt_id
        le.reason = reason
        try:
            setattr(le, 'approver', approver or None)
        except Exception:
            pass
        try:
            setattr(le, 'recorder_name', recorder_name or None)   
        except Exception:
            pass
        db.session.commit()

        lt = LeaveType.query.get(lt_id)
        is_paid = bool(lt and getattr(lt, 'is_paid', True))

        tr = Transaction.query.filter_by(reference_id=le.id, type='LEAVE_TAKEN').first()
        if tr:
            tr.amount = round(-abs(days),2) if is_paid else 0.0
            note_txt = f'Edited leave {date_from} to {date_to}'
            if approver:
                note_txt += f' — Approver: {approver}'
            situation = request.form.get('situation','').strip() or None
            le.situation = situation
            if not is_paid:
                note_txt += ' (UNPAID – no balance deduction)'
            tr.note = note_txt
            db.session.commit()

        flash('Leave updated', 'success')
        return redirect(url_for('employee_detail', emp_id=emp.id))

    return render_template('edit_leave.html', le=le, leave_types=leave_types, emp=emp) # type: ignore

@app.route("/leave/delete/<int:leave_id>", methods=["POST"])
@login_required
def delete_leave(leave_id):
    if not has_permission("can_delete_leave") and session.get("role") not in (
        "developer",
        "admin_master",
    ):
        flash("Delete permission required", "danger")
        return redirect(url_for("index"))
    le = LeaveEntry.query.get_or_404(leave_id)
    emp = Employee.query.get_or_404(le.employee_id)

    emp = Employee.query.get_or_404(le.employee_id)
    # Block deleting leave for left employees (dev only)
    if is_employee_left(emp) and session.get("role") != "developer":
        flash("Leave records of a left employee cannot be deleted.", "danger")
        return redirect(url_for("employee_detail", emp_id=emp.id))

    tr_list = Transaction.query.filter_by(reference_id=le.id).all()
    for tr in tr_list:
        db.session.delete(tr)
    db.session.delete(le)
    adj = Transaction(
        employee_id=emp.id,
        type="ADJUSTMENT",
        period=None,
        amount=0.0,
        note=f"Deleted leave id {leave_id} by {session.get('username')}",
        created_by=session.get("user_id"),
    )
    db.session.add(adj)
    db.session.commit()
    flash("Leave deleted", "success")
    return redirect(url_for("employee_detail", emp_id=emp.id))


# ---------- Employee detail, print & export ----------
@app.route("/employee/<int:emp_id>")
@login_required
def employee_detail(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    bal = compute_balance(emp)
    leaves = (
        LeaveEntry.query.filter_by(employee_id=emp.id)
        .order_by(LeaveEntry.date_from.desc())
        .all()
    )
    return render_template(
    "employee_detail.html", e=emp, bal=bal, leaves=leaves
    )   

@app.route("/employee/print/<int:emp_id>")
@login_required
def employee_print(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    bal = compute_balance(emp)
    leaves = (
        LeaveEntry.query.filter_by(employee_id=emp.id)
        .order_by(LeaveEntry.date_from)
        .all()
    )
    transactions = Transaction.query.filter(
        Transaction.employee_id == emp.id,
        Transaction.type.in_(['LEAVE_TAKEN'])
    ).order_by(Transaction.created_at).all()

    return render_template(
        "employee_print.html", e=emp, bal=bal, leaves=leaves, transactions=transactions
    )

@app.route("/employee/<int:emp_id>/export")
@login_required
def export_employee_profile(emp_id):
    emp = Employee.query.get_or_404(emp_id)

    # ----------------------------
    # Collect data for this employee ONLY
    # ----------------------------
    leaves = LeaveEntry.query.filter_by(employee_id=emp.id)\
        .order_by(LeaveEntry.date_from).all()

    compoffs = CompOffRecord.query.filter_by(employee_id=emp.id)\
        .order_by(CompOffRecord.earned_on).all()

    early_lates = EarlyLateRecord.query.filter_by(employee_id=emp.id)\
        .order_by(EarlyLateRecord.created_at).all()

    ods = OutdoorDuty.query.filter_by(employee_id=emp.id)\
        .order_by(OutdoorDuty.od_date).all()

    wb = Workbook()

    # =====================================================
    # SHEET 1 — EMPLOYEE PROFILE
    # =====================================================
    ws = wb.active
    ws.title = "Profile"

    ws.append(["Field", "Value"])
    profile_rows = [
        ("Employee ID", emp.employee_id),
        ("Full Name", f"{emp.first_name} {emp.middle_name or ''} {emp.last_name}".strip()),
        ("Department", emp.department or ""),
        ("Designation", emp.designation or ""),
        ("Plant Location", emp.plant_location or ""),
        ("Contact Number", emp.contact_number or ""),
        ("Emergency Number", emp.emergency_number or ""),
        ("Hire Date", emp.hire_date.strftime("%d/%m/%Y") if emp.hire_date else ""),
        ("Status", emp.status),
        ("Left Date", emp.left_date.strftime("%d/%m/%Y") if emp.left_date else ""),
        ("Accrual Rate", emp.accrual_rate),
    ]

    for r in profile_rows:
        ws.append(r)

    # =====================================================
    # SHEET 2 — LEAVES
    # =====================================================
    ws = wb.create_sheet("Leaves")
    ws.append([
        "From Date", "To Date", "Days",
        "Leave Type", "Situation",
        "Approver", "Recorded By", "Reason"
    ])

    for l in leaves:
        ws.append([
            l.date_from.strftime("%d/%m/%Y") if l.date_from else "",
            l.date_to.strftime("%d/%m/%Y") if l.date_to else "",
            l.days,
            l.leave_type.name if l.leave_type else "",
            l.situation or "",
            l.approver or "",
            l.recorder_name or "",
            l.reason or "",
        ])

    # =====================================================
    # SHEET 3 — COMP OFF
    # =====================================================
    ws = wb.create_sheet("Comp-Off")
    ws.append([
        "Earned On", "Taken On",
        "Approved By", "Note"
    ])

    for c in compoffs:
        ws.append([
            c.earned_on.strftime("%d/%m/%Y") if c.earned_on else "",
            c.taken_on.strftime("%d/%m/%Y") if c.taken_on else "",
            c.approved_by or "",
            c.note or "",
        ])

    # =====================================================
    # SHEET 4 — EARLY / LATE
    # =====================================================
    ws = wb.create_sheet("Early-Late")
    ws.append([
        "Date", "Late Time", "Early Time", "Note"
    ])

    for r in early_lates:
        date_val = (
            r.late_datetime or r.early_datetime
        )
        ws.append([
            date_val.strftime("%d/%m/%Y") if date_val else "",
            r.late_datetime.strftime("%H:%M") if r.late_datetime else "",
            r.early_datetime.strftime("%H:%M") if r.early_datetime else "",
            r.note or "",
        ])

    # =====================================================
    # SHEET 5 — OUTDOOR DUTY
    # =====================================================
    ws = wb.create_sheet("Outdoor Duty")
    ws.append([
        "Date", "Full Day",
        "Time From", "Time To",
        "Approved By", "Reason", "Note"
    ])

    for o in ods:
        ws.append([
            o.od_date.strftime("%d/%m/%Y") if o.od_date else "",
            "Yes" if o.is_full_day else "No",
            o.time_from.strftime("%H:%M") if o.time_from else "",
            o.time_to.strftime("%H:%M") if o.time_to else "",
            o.approved_by or "",
            o.reason or "",
            o.note or "",
        ])

    # ----------------------------
    # Auto-size columns
    # ----------------------------
    for ws in wb.worksheets:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = (
                max(len(str(c.value or "")) for c in col) + 3
            )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{emp.employee_id}_profile.xlsx"
    )

@app.route("/check_balance", methods=["GET", "POST"])
@login_required
def check_balance():
    """
    Show a small form to lookup an employee by Employee ID and display current balance + recent leaves.
    """
    if request.method == "POST":
        emp_code = request.form.get("employee_id", "").strip()
        emp = Employee.query.filter_by(employee_id=emp_code).first()
        if not emp:
            flash("Employee not found", "danger")
            return redirect(url_for("check_balance"))
        bal = compute_balance(emp)
        leaves = (
            LeaveEntry.query.filter_by(employee_id=emp.id)
            .order_by(LeaveEntry.date_from.desc())
            .all()
        )
        return render_template("balance.html", emp=emp, bal=bal, leaves=leaves)

    return render_template("check_balance.html")


@app.route("/export_employees")
@login_required
def export_employees():
    """
    Export employees list.
    Default = CSV
    ?format=xlsx -> Excel
    """
    status = request.args.get("status", "")
    q = request.args.get("q", "").strip()
    fmt = request.args.get("format", "").lower()

    base = Employee.query
    if status:
        base = base.filter(Employee.status == status)
    if q:
        base = base.filter(
            or_(
                Employee.employee_id.ilike(f"%{q}%"),
                Employee.first_name.ilike(f"%{q}%"),
                Employee.last_name.ilike(f"%{q}%"),
            )
        )

    employees = base.order_by(Employee.employee_id).all()

    # ================= XLSX =================
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "Employee ID", "First name", "Middle name", "Last name",
            "Plant Location", "Department", "Designation",
            "Contact", "Emergency",
            "Hire Date", "Status", "Left On",
            "Accrual rate", "Manual balance"
        ]

        ws.append(headers)

        left_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )

        for e in employees:
            is_left = is_employee_left(e)
            left_on = ""
            if e.left_date:
                try:
                    left_on = e.left_date.isoformat()
                except Exception:
                    left_on = str(e.left_date)

            row = [
                e.employee_id,
                e.first_name,
                e.middle_name or "",
                e.last_name,
                e.plant_location or "",
                e.department or "",
                e.designation or "",
                e.contact_number or "",
                e.emergency_number or "",
                e.hire_date.isoformat() if e.hire_date else "",
                "left" if is_left else "active",
                left_on,
                e.accrual_rate,
                getattr(e, "manual_balance", ""),
            ]

            ws.append(row)

            if is_left:
                for cell in ws[ws.max_row]:
                    cell.fill = left_fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value or "")) for c in col
            ) + 2

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="employees.xlsx",
        )

    # ================= CSV (DEFAULT) =================
    si = StringIO()
    writer = csv.writer(si)

    writer.writerow([
        "Employee ID", "First name", "Middle name", "Last name",
        "Plant Location", "Department", "Designation",
        "Contact", "Emergency",
        "Hire Date", "Status", "Left On",
        "Accrual rate", "Manual balance"
    ])

    for e in employees:
        is_left = is_employee_left(e)
        left_on = e.left_date.isoformat() if e.left_date else ""

        writer.writerow([
            e.employee_id,
            e.first_name,
            e.middle_name or "",
            e.last_name,
            e.plant_location or "",
            e.department or "",
            e.designation or "",
            e.contact_number or "",
            e.emergency_number or "",
            e.hire_date.isoformat() if e.hire_date else "",
            "left" if is_left else "active",
            left_on,
            e.accrual_rate,
            getattr(e, "manual_balance", ""),
        ])

    output = si.getvalue().encode("utf-8")
    return send_file(
        BytesIO(output),
        mimetype="text/csv",
        as_attachment=True,
        download_name="employees.csv",
    )

# ---------- History & exports ----------
@app.route("/history")
@login_required
def history_home():
    return redirect(url_for("history_leaves"))

@app.route('/history/leaves', methods=['GET'])
@login_required
def history_leaves():
    emp_code = request.args.get('employee_id','').strip()
    leave_type_id = request.args.get('leave_type_id','')
    dept = request.args.get('department','')
    situation = request.args.get('situation','').strip()

    today = date.today()
    first_day = date(today.year, today.month, 1)

    if today.month == 12:
        last_day = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(today.year, today.month + 1, 1) - timedelta(days=1)

    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    emp_code = request.args.get('employee_id', '').strip()
    leave_type_id = request.args.get('leave_type_id', '')
    dept = request.args.get('department', '')
    situation = request.args.get('situation', '').strip()

    q = LeaveEntry.query.filter(
        LeaveEntry.date_from <= last_day,
        LeaveEntry.date_to >= first_day
    )

    if from_date:
        try:
            fd = dt.strptime(from_date, '%Y-%m-%d').date()
            q = q.filter(LeaveEntry.date_from >= fd)
        except Exception:
            pass
    if to_date:
        try:
            td = dt.strptime(to_date, '%Y-%m-%d').date()
            q = q.filter(LeaveEntry.date_to <= td)
        except Exception:
            pass

    if emp_code:
        emp = Employee.query.filter_by(employee_id=emp_code).first()
        q = q.filter(LeaveEntry.employee_id == emp.id) if emp else q.filter(LeaveEntry.employee_id == -1)

    if leave_type_id:
        try:
            q = q.filter(LeaveEntry.leave_type_id == int(leave_type_id))
        except Exception:
            pass

    if dept:
        emp_ids = [e.id for e in Employee.query.filter(Employee.department == dept).all()]
        q = q.filter(LeaveEntry.employee_id.in_(emp_ids)) if emp_ids else q.filter(LeaveEntry.employee_id == -1)

    if situation:
        q = q.filter(LeaveEntry.situation == situation)

    leaves = q.order_by(LeaveEntry.date_from.desc()).limit(2000).all()

    emp_ids = {l.employee_id for l in leaves}
    employees = Employee.query.filter(Employee.id.in_(list(emp_ids))).all() if emp_ids else []
    emp_map = {e.id: e for e in employees}

    current_year = dt.now().year

    users = {u.id: u.username for u in User.query.all()}

    return render_template(
        'leaves_history.html',
        leaves=leaves,
        emp_map=emp_map,
        leave_types=LeaveType.query.all(),
        departments=['Purchase','ME','HR Admin','Production','Quality','Store','Other'],
        situations=['Planned','Unplanned','Sick'],
        filters={
            'from_date': from_date,
            'to_date': to_date,
            'employee_id': emp_code,
            'leave_type_id': leave_type_id,
            'department': dept,
            'situation': situation
        },
        active_tab='leaves'
    )

@app.route('/history/compoff')
@login_required
def history_compoff():
    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()
    q_year = request.args.get('year','').strip()

    # 🔒 FORCE CURRENT MONTH ONLY
    today = date.today()
    first_day = date(today.year, today.month, 1)
    if today.month == 12:
        last_day = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(today.year, today.month + 1, 1) - timedelta(days=1)

    q = CompOffRecord.query.filter(
        CompOffRecord.earned_on >= first_day,
        CompOffRecord.earned_on <= last_day
    ).order_by(CompOffRecord.earned_on.desc())

    if q_emp:
        q = q.filter(CompOffRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(CompOffRecord.department == q_dept)

    items = q.limit(2000).all()

    users = {u.id: u.username for u in User.query.all()}

    return render_template(
        'compoff_history.html',
        items=items,
        active_tab='compoff',
        filters={'employee_id': q_emp, 'department': q_dept, 'year': q_year}
    )

@app.route('/history/early-late')
@login_required
def history_early_late():
    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()
    q_year = request.args.get('year','').strip()
    q_month = request.args.get('month','').strip()

    today = date.today()
    current_year = today.year
    current_month = today.month

    q = EarlyLateRecord.query.filter(
        or_(
            and_(
                EarlyLateRecord.late_datetime.isnot(None),
                db.extract('year', EarlyLateRecord.late_datetime) == current_year,
                db.extract('month', EarlyLateRecord.late_datetime) == current_month
            ),
            and_(
                EarlyLateRecord.early_datetime.isnot(None),
                db.extract('year', EarlyLateRecord.early_datetime) == current_year,
                db.extract('month', EarlyLateRecord.early_datetime) == current_month
            )
        )
    ).order_by(
        EarlyLateRecord.late_datetime.desc().nullslast(),
        EarlyLateRecord.early_datetime.desc().nullslast()
    )

    if q_emp:
        q = q.filter(EarlyLateRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(EarlyLateRecord.department == q_dept)
    if q_year:
        try:
            y = int(q_year)
            q = q.filter(
                or_(
                    db.extract('year', EarlyLateRecord.late_datetime) == y,
                    db.extract('year', EarlyLateRecord.early_datetime) == y
                )
            )
        except Exception:
            pass


    if q_month:
        try:
            m = int(q_month)
            q = q.filter(
                or_(
                    db.extract('month', EarlyLateRecord.late_datetime) == m,
                    db.extract('month', EarlyLateRecord.early_datetime) == m
                )
            )
        except Exception:
            pass

    items = q.limit(2000).all()

    users = {u.id: u.username for u in User.query.all()}

    return render_template(
        'early_late_list.html',
        items=items,
        pagination=None,
        q_emp=q_emp,
        q_dept=q_dept,
        q_year=q_year,
        q_month=q_month,
        active_tab='early_late'
    )

@app.route('/history/outdoor')
@login_required
def history_outdoor():

    q = OutdoorDuty.query.order_by(OutdoorDuty.od_date.desc())

    items = q.limit(2000).all()

    filters = {
        "employee_id": "",
        "department": "",
        "year": "",
        "month": ""
    }

    users = {u.id: u.username for u in User.query.all()}

    return render_template(
        'outdoor_duty_history.html',
        items=items,
        filters=filters,
        active_tab='outdoor'
    )

@app.route('/history/leaves/export')
@login_required
def export_history_leaves():
    today = date.today()
    first_day = date(today.year, today.month, 1)
    last_day = (
        date(today.year + 1, 1, 1) - timedelta(days=1)
        if today.month == 12
        else date(today.year, today.month + 1, 1) - timedelta(days=1)
    )

    emp_code = request.args.get('employee_id','').strip()
    leave_type_id = request.args.get('leave_type_id','')
    dept = request.args.get('department','')
    situation = request.args.get('situation','').strip()

    q = LeaveEntry.query.filter(
        LeaveEntry.date_from <= last_day,
        LeaveEntry.date_to >= first_day
    )

    if emp_code:
        emp = Employee.query.filter_by(employee_id=emp_code).first()
        q = q.filter(LeaveEntry.employee_id == emp.id) if emp else q.filter(LeaveEntry.employee_id == -1)

    if leave_type_id:
        try:
            q = q.filter(LeaveEntry.leave_type_id == int(leave_type_id))
        except Exception:
            pass

    if dept:
        emp_ids = [e.id for e in Employee.query.filter(Employee.department == dept).all()]
        q = q.filter(LeaveEntry.employee_id.in_(emp_ids)) if emp_ids else q.filter(LeaveEntry.employee_id == -1)

    if situation:
        q = q.filter(LeaveEntry.situation == situation)

    leaves = q.order_by(LeaveEntry.date_from).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Leaves_{today.strftime('%b_%Y')}"

    ws.append([
        'Employee ID', 'Name', 'Department',
        'Leave Type', 'Situation',
        'From Date', 'To Date', 'Days',
        'Approver', 'Recorded By', 'Reason'
    ])

    emp_ids = {l.employee_id for l in leaves}
    emp_map = {
        e.id: e for e in Employee.query.filter(Employee.id.in_(emp_ids)).all()
    } if emp_ids else {}

    lt_map = {
        lt.id: lt.name for lt in LeaveType.query.all()
    }

    for l in leaves:
        emp = emp_map.get(l.employee_id)
        ws.append([
            emp.employee_id if emp else '',
            f"{emp.first_name} {emp.last_name}".strip() if emp else '',
            emp.department if emp else '',
            lt_map.get(l.leave_type_id, ''),
            l.situation or '',
            l.date_from.strftime('%d/%m/%Y') if l.date_from else '',
            l.date_to.strftime('%d/%m/%Y') if l.date_to else '',
            l.days,
            l.approver or '',
            l.recorder_name or '',
            l.reason or ''
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or '')) for c in col
        ) + 3

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'leave_history_{today.strftime("%Y_%m")}.xlsx'
    )

@app.route("/month_report")
@login_required
def month_report():
    """
    Small UI to pick year/month before exporting the month report.
    """
    return render_template("month_report.html")

@app.route('/export_month_report', methods=['GET'])
@login_required
def export_month_report():
    try:
        year_i = int(request.args.get('year'))
        month_i = int(request.args.get('month'))
    except Exception:
        flash('Please select both Year and Month.', 'danger')
        return redirect(url_for('month_report'))

    first_day = date(year_i, month_i, 1)
    last_day = (
        date(year_i + 1, 1, 1) - timedelta(days=1)
        if month_i == 12
        else date(year_i, month_i + 1, 1) - timedelta(days=1)
    )

    wb = Workbook()
    wb.remove(wb.active)  

    # =====================================================
    # 1️⃣ LEAVES SHEET
    # =====================================================
    leaves = LeaveEntry.query.filter(
        LeaveEntry.date_to >= first_day,
        LeaveEntry.date_from <= last_day
    ).order_by(LeaveEntry.date_from).all()

    ws = wb.create_sheet(f'Leaves_{year_i}_{str(month_i).zfill(2)}')
    ws.append([
        'Employee ID', 'Name', 'Department',
        'Leave Type', 'Situation',
        'From Date', 'To Date', 'Days',
        'Approver', 'Recorded By', 'Reason'
    ])

    emp_ids = {l.employee_id for l in leaves}
    emp_map = {
        e.id: e for e in Employee.query.filter(Employee.id.in_(emp_ids)).all()
    } if emp_ids else {}

    lt_ids = {l.leave_type_id for l in leaves if l.leave_type_id}
    lt_map = {
        lt.id: lt.name for lt in LeaveType.query.filter(LeaveType.id.in_(lt_ids)).all()
    } if lt_ids else {}

    for l in leaves:
        emp = emp_map.get(l.employee_id)
        ws.append([
            emp.employee_id if emp else '',
            f"{emp.first_name} {emp.last_name}".strip() if emp else '',
            emp.department if emp else '',
            lt_map.get(l.leave_type_id, ''),
            l.situation or '',
            l.date_from.strftime('%d/%m/%Y') if l.date_from else '',
            l.date_to.strftime('%d/%m/%Y') if l.date_to else '',
            l.days,
            l.approver or '',
            l.recorder_name or '',
            l.reason or ''
        ])

    # =====================================================
    # 2️⃣ COMP-OFF SHEET (NO "Recorded By")
    # =====================================================
    compoffs = CompOffRecord.query.filter(
        CompOffRecord.earned_on.between(first_day, last_day)
    ).order_by(CompOffRecord.earned_on).all()

    ws = wb.create_sheet(f'CompOff_{year_i}_{str(month_i).zfill(2)}')
    ws.append([
        'Employee ID', 'Name', 'Department',
        'Earned On', 'Taken On',
        'Approved By', 'Note'
    ])

    for c in compoffs:
        ws.append([
            c.emp_code,
            c.emp_name,
            c.department or '',
            c.earned_on.strftime('%d/%m/%Y') if c.earned_on else '',
            c.taken_on.strftime('%d/%m/%Y') if c.taken_on else '',
            c.approved_by or '',
            c.note or ''
        ])

    # =====================================================
    # 3️⃣ EARLY / LATE SHEET (NO Recorded / Approved By)
    # =====================================================
    early_late = EarlyLateRecord.query.filter(
        or_(
            EarlyLateRecord.late_datetime.between(first_day, last_day),
            EarlyLateRecord.early_datetime.between(first_day, last_day)
        )
    ).order_by(EarlyLateRecord.created_at).all()

    ws = wb.create_sheet(f'EarlyLate_{year_i}_{str(month_i).zfill(2)}')
    ws.append([
        'Employee ID', 'Name', 'Department',
        'Date', 'Late Time', 'Early Time', 'Note'
    ])

    for r in early_late:
        event_date = (
            r.late_datetime.date() if r.late_datetime
            else r.early_datetime.date() if r.early_datetime
            else None
        )
        ws.append([
            r.emp_code,
            r.emp_name,
            r.department or '',
            event_date.strftime('%d/%m/%Y') if event_date else '',
            r.late_datetime.strftime('%H:%M') if r.late_datetime else '',
            r.early_datetime.strftime('%H:%M') if r.early_datetime else '',
            r.note or ''
        ])

    # =====================================================
    # 4️⃣ OUTDOOR DUTY SHEET
    # =====================================================
    ods = OutdoorDuty.query.filter(
        OutdoorDuty.od_date.between(first_day, last_day)
    ).order_by(OutdoorDuty.od_date).all()

    ws = wb.create_sheet(f'OD_{year_i}_{str(month_i).zfill(2)}')
    ws.append([
        'Employee ID', 'Name', 'Department', 'Designation',
        'OD Date', 'Full Day',
        'From Time', 'To Time',
        'Approved By', 'Reason', 'Note'
    ])

    for o in ods:
        ws.append([
            o.emp_code,
            o.emp_name,
            o.department or '',
            o.designation or '',
            o.od_date.strftime('%d/%m/%Y') if o.od_date else '',
            'Yes' if o.is_full_day else 'No',
            o.time_from.strftime('%H:%M') if o.time_from else '',
            o.time_to.strftime('%H:%M') if o.time_to else '',
            o.approved_by or '',
            o.reason or '',
            o.note or ''
        ])

    # =====================================================
    # AUTO-SIZE ALL SHEETS
    # =====================================================
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'monthly_report_{year_i}_{str(month_i).zfill(2)}.xlsx'
    )

@app.route('/outdoor/export')
@login_required
def outdoor_export():

    rows = []

    for r in OutdoorDuty.query.order_by(OutdoorDuty.od_date.asc()).all():

        if r.is_full_day:
            od_type = 'Full Day'
            time_from = ''
            time_to = ''
        else:
            od_type = 'Half Day'
            time_from = r.time_from.strftime('%H:%M') if r.time_from else ''
            time_to = r.time_to.strftime('%H:%M') if r.time_to else ''

        rows.append({
            'EMPLOYEE_ID': r.emp_code,
            'NAME': r.emp_name,
            'DEPARTMENT': r.department or '',
            'DESIGNATION': r.designation or '',
            'OD_DATE': r.od_date.strftime('%d-%m-%Y'),
            'OD_TYPE': od_type,
            'FROM_TIME': time_from,
            'TO_TIME': time_to,
            'REASON': r.reason or '',
            'APPROVED_BY': r.approved_by or '',
            'NOTE': r.note or ''
        })

    df = pd.DataFrame(rows)

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='OutdoorDuty')

    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name='outdoor_duty.xlsx',
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/internal/ledger')
@login_required
def internal_ledger():
    # 🔒 DEV / ADMIN MASTER ONLY
    if session.get("role") not in ("developer"):
        flash("Access denied.", "danger")
        return redirect(url_for("index"))

    q_emp = request.args.get('employee_id','').strip()
    q_type = request.args.get('type','').strip()
    q_year = request.args.get('year','').strip()

    q = Transaction.query.order_by(Transaction.created_at.desc())

    if q_emp:
        emp = Employee.query.filter_by(employee_id=q_emp).first()
        q = q.filter(Transaction.employee_id == emp.id) if emp else q.filter(Transaction.employee_id == -1)

    if q_type:
        q = q.filter(Transaction.type == q_type)

    if q_year:
        try:
            q = q.filter(db.extract('year', Transaction.created_at) == int(q_year))
        except Exception:
            pass

    items = q.limit(3000).all()

    emp_ids = {t.employee_id for t in items if t.employee_id}
    employees = Employee.query.filter(Employee.id.in_(list(emp_ids))).all() if emp_ids else []
    emp_map = {e.id: e for e in employees}

    return render_template(
        'internal_ledger.html',
        items=items,
        emp_map=emp_map,
        filters={
            'employee_id': q_emp,
            'type': q_type,
            'year': q_year
        }
    )

@app.route('/export_year_report', methods=['GET'])
@login_required
def export_year_report():
    year = request.args.get('year')
    try:
        year_i = int(year) if year else date.today().year
    except Exception:
        year_i = date.today().year

    start_date = date(year_i, 1, 1)
    end_date = date(year_i, 12, 31)

    wb = Workbook()

    # =========================================================
    # SHEET 1 — LEAVES (Recorded By = recorder_name ONLY)
    # =========================================================
    ws_leaves = wb.active
    ws_leaves.title = 'Leaves'

    ws_leaves.append([
        'Employee ID', 'Name', 'Department',
        'Leave Type', 'Situation',
        'From Date', 'To Date', 'Days',
        'Approver', 'Recorded By', 'Reason'
    ])

    leaves = LeaveEntry.query.filter(
        LeaveEntry.date_to >= start_date,
        LeaveEntry.date_from <= end_date
    ).order_by(LeaveEntry.date_from).all()

    emp_ids = {l.employee_id for l in leaves}
    emp_map = {
        e.id: e for e in Employee.query.filter(Employee.id.in_(emp_ids)).all()
    } if emp_ids else {}

    lt_ids = {l.leave_type_id for l in leaves if l.leave_type_id}
    lt_map = {
        lt.id: lt.name for lt in LeaveType.query.filter(LeaveType.id.in_(lt_ids)).all()
    } if lt_ids else {}

    for l in leaves:
        emp = emp_map.get(l.employee_id)
        ws_leaves.append([
            emp.employee_id if emp else '',
            f"{emp.first_name} {emp.last_name}".strip() if emp else '',
            emp.department if emp else '',
            lt_map.get(l.leave_type_id, ''),
            l.situation or '',
            l.date_from.strftime('%d/%m/%Y') if l.date_from else '',
            l.date_to.strftime('%d/%m/%Y') if l.date_to else '',
            l.days,
            l.approver or '',
            l.recorder_name or '',
            l.reason or ''
        ])

    # =========================================================
    # SHEET 2 — COMP-OFF (NO Recorded By column)
    # =========================================================
    ws_co = wb.create_sheet('Comp-Off')
    ws_co.append([
        'Employee ID', 'Name', 'Department',
        'Earned On', 'Taken On',
        'Approved By', 'Note'
    ])

    compoffs = CompOffRecord.query.filter(
        CompOffRecord.earned_on.between(start_date, end_date)
    ).order_by(CompOffRecord.earned_on).all()

    for c in compoffs:
        ws_co.append([
            c.emp_code,
            c.emp_name,
            c.department or '',
            c.earned_on.strftime('%d/%m/%Y') if c.earned_on else '',
            c.taken_on.strftime('%d/%m/%Y') if c.taken_on else '',
            c.approved_by or '',
            c.note or ''
        ])

    # =========================================================
    # SHEET 3 — EARLY / LATE (NO Recorded By / Approved By)
    # =========================================================
    ws_el = wb.create_sheet('Early-Late')
    ws_el.append([
        'Employee ID', 'Name', 'Department',
        'Date', 'Late Time', 'Early Time', 'Note'
    ])

    el_records = EarlyLateRecord.query.filter(
        or_(
            EarlyLateRecord.late_datetime.between(start_date, end_date),
            EarlyLateRecord.early_datetime.between(start_date, end_date)
        )
    ).order_by(EarlyLateRecord.late_datetime, EarlyLateRecord.early_datetime).all()

    for r in el_records:
        dt_val = r.late_datetime or r.early_datetime
        ws_el.append([
            r.emp_code,
            r.emp_name,
            r.department or '',
            dt_val.strftime('%d/%m/%Y') if dt_val else '',
            r.late_datetime.strftime('%H:%M') if r.late_datetime else '',
            r.early_datetime.strftime('%H:%M') if r.early_datetime else '',
            r.note or ''
        ])

    # =========================================================
    # SHEET 4 — OUTDOOR DUTY (Approved By ONLY)
    # =========================================================
    ws_od = wb.create_sheet('Outdoor-Duty')
    ws_od.append([
        'Employee ID', 'Name', 'Department',
        'Date', 'Type',
        'From Time', 'To Time',
        'Approved By', 'Reason', 'Note'
    ])

    ods = OutdoorDuty.query.filter(
        OutdoorDuty.od_date.between(start_date, end_date)
    ).order_by(OutdoorDuty.od_date).all()

    for o in ods:
        ws_od.append([
            o.emp_code,
            o.emp_name,
            o.department or '',
            o.od_date.strftime('%d/%m/%Y') if o.od_date else '',
            'Full Day' if o.is_full_day else 'Half Day',
            o.time_from.strftime('%H:%M') if o.time_from else '',
            o.time_to.strftime('%H:%M') if o.time_to else '',
            o.approved_by or '',
            o.reason or '',
            o.note or ''
        ])

    # =========================================================
    # AUTO SIZE ALL SHEETS
    # =========================================================
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'year_report_{year_i}.xlsx'
    ) 

# ---------- Backup & Restore ----------
@app.route("/backup")
@login_required
def backup():

    print("BACKUP ROLE =", repr(session.get("role")))

    role = normalize_role(session.get("role"))

    if role not in ("developer", "admin_master", "admin_1"):
        flash("You are not allowed to create backups.", "danger")
        return redirect(url_for("index"))

    if not os.path.exists(DB_PATH) or os.path.getsize(DB_PATH) == 0:
        flash("Database not found or empty.", "danger")
        return redirect(url_for("index"))

    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"app_backup_{timestamp}.db"
    backup_path = os.path.join(BACKUPS_FOLDER, backup_name)

    try:
        shutil.copy2(DB_PATH, backup_path)
    except Exception as e:
        flash(f"Failed to create backup: {e}", "danger")
        return redirect(url_for("index"))

    return send_file(
        backup_path,
        as_attachment=True,
        download_name=backup_name
    )

@app.route("/restore", methods=["GET", "POST"])
@login_required
def restore():
    """
    Restore the DB from an uploaded file.
    Only developer or admin_master allowed.
    """
    role = normalize_role(session.get("role"))

    if role not in ("developer", "admin_master"):
        flash("Restore requires admin_master or developer role", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        if "db_file" not in request.files:
            flash("No file uploaded", "danger")
            return redirect(url_for("restore"))
        f = request.files["db_file"]
        if f.filename == "":
            flash("No selected file", "danger")
            return redirect(url_for("restore"))
        filename = secure_filename(f.filename)
        dest = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        f.save(dest)

        try:
            if os.path.exists(DB_PATH):
                pre_bak = os.path.join(
                    BACKUPS_FOLDER,
                    "pre_restore_" + dt.now().strftime("%Y%m%d_%H%M%S") + ".db",
                )
                shutil.copy2(DB_PATH, pre_bak)
            shutil.copy2(dest, DB_PATH)
        except Exception as e:
            flash(f"Failed to restore DB: {e}", "danger")
            return redirect(url_for("restore"))

        flash(
            "Database restored. Please restart the application to ensure a clean state.",
            "success",
        )
        return redirect(url_for("index"))

    return render_template("restore.html")

# ----------------------------
# Employee import (xlsx)
# ----------------------------
IMPORT_HEADERS = [
    "employee_id","first_name","middle_name","last_name","hire_date","accrual_rate",
    "plant_location","department","designation","contact_number","emergency_number",
    "status","manual_balance"
]

def _normalize_header(h):
    return str(h).strip().lower().replace(' ', '_')

def is_valid_phone(n):
    return isinstance(n, str) and n.isdigit() and len(n) == 10

def parse_xlsx_file(path):
    """
    Parse .xlsx and return list of row dicts and a list of file-level errors.
    Each row dict: {'row': int, 'data': {col: val,...}, 'errors': [str,...]}
    """
    rows = []
    errors = []
    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
    except Exception as e:
        return [], [f'Failed to open file: {e}']

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if any(cell is not None and str(cell).strip() != '' for cell in row):
            header_row = row
            break
    if not header_row:
        return [], ['No header row found']

    headers = [ _normalize_header(h) if h is not None else '' for h in header_row ]
    col_map = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[idx] = h

    required_headers = ['employee_id','first_name','last_name','hire_date','accrual_rate']
    for rh in required_headers:
        if rh not in headers:
            errors.append(f"Missing required column: {rh}")
    if errors:
        return [], errors

    seen_ids_in_file = set()
    row_index = (ws.min_row if ws.min_row else 2)  
    header_row_num = None
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [c for c in r]
        normalized = [ _normalize_header(c) if c is not None else '' for c in cells ]
        if 'employee_id' in normalized and 'first_name' in normalized:
            header_row_num = i
            headers = normalized
            break
    if not header_row_num:
        return [], ['Failed to detect header row (expected employee_id, first_name)']

    for i, row in enumerate(ws.iter_rows(min_row=header_row_num+1, values_only=True), start=header_row_num+1):
        row_dict = {}
        for idx, val in enumerate(row):
            h = headers[idx] if idx < len(headers) else None
            if not h:
                continue
            if isinstance(val, str):
                v = val.strip()
            else:
                v = val
            row_dict[h] = v
        data = {}
        for key in IMPORT_HEADERS:
            data[key] = row_dict.get(key, None)
        row_errors = []
        contact = data.get('contact_number')

        if not contact or str(contact).strip() == '':
            row_errors.append('contact_number is required')
        else:
            contact = str(contact).strip()
            if not is_valid_phone(contact):
                row_errors.append('contact_number must be exactly 10 digits')
            elif Employee.query.filter_by(contact_number=contact).first():
                row_errors.append('contact_number already exists in database')

        emergency = data.get('emergency_number')
        if emergency not in (None, ''):
            emergency = str(emergency).strip()
            if not is_valid_phone(emergency):
                row_errors.append('emergency_number must be exactly 10 digits')

        empid = (data.get('employee_id') or '')
        if not empid or str(empid).strip() == '':
            row_errors.append('employee_id is required')
        else:
            empid = str(empid).strip()
            if empid in seen_ids_in_file:
                row_errors.append('duplicate employee_id in file')
            else:
                seen_ids_in_file.add(empid)
        if not data.get('first_name') or str(data.get('first_name')).strip() == '':
            row_errors.append('first_name is required')
        if not data.get('last_name') or str(data.get('last_name')).strip() == '':
            row_errors.append('last_name is required')
        hd = data.get('hire_date')
        parsed_hire = None
        if hd in (None, ''):
            row_errors.append('hire_date is required')
        else:
            if isinstance(hd, (str,)):
                try:
                    parsed_hire = dt.strptime(hd.strip(), '%Y-%m-%d').date()
                except Exception:
                    try:
                        parsed_hire = dt.strptime(hd.strip(), '%d-%m-%Y').date()
                    except Exception:
                        row_errors.append('hire_date must be YYYY-MM-DD or DD-MM-YYYY')
            elif isinstance(hd, (dt, )):
                parsed_hire = hd.date() if isinstance(hd, dt) else None
            elif hasattr(hd, 'year') and hasattr(hd, 'month'):
                try:
                    parsed_hire = hd
                except Exception:
                    parsed_hire = None
            else:
                row_errors.append('Invalid hire_date format')

        accr = data.get('accrual_rate')
        parsed_accr = None
        try:
            parsed_accr = float(accr)
            if parsed_accr <= 0:
                row_errors.append('accrual_rate must be positive')
        except Exception:
            row_errors.append('accrual_rate must be numeric')

        status = (data.get('status') or 'active')
        if isinstance(status, str):
            status_norm = status.strip().lower()
            if status_norm not in ('active','left',''):
                row_errors.append('status must be active or left (or blank)')
            if status_norm == '':
                status_norm = 'active'
        else:
            status_norm = 'active'

        manual_balance = data.get('manual_balance')
        parsed_manual = None
        if manual_balance not in (None, ''):
            try:
                parsed_manual = float(manual_balance)
            except Exception:
                row_errors.append('manual_balance must be numeric if provided')

        cleaned = {
            'employee_id': str(empid).strip() if empid else '',
            'first_name': str(data.get('first_name') or '').strip(),
            'middle_name': (str(data.get('middle_name')).strip() if data.get('middle_name') else None),
            'last_name': str(data.get('last_name') or '').strip(),
            'hire_date': parsed_hire,
            'accrual_rate': parsed_accr,
            'plant_location': (str(data.get('plant_location')).strip() if data.get('plant_location') else None),
            'department': (str(data.get('department')).strip() if data.get('department') else None),
            'designation': (str(data.get('designation')).strip() if data.get('designation') else None),
            'contact_number': (str(data.get('contact_number')).strip() if data.get('contact_number') else None),
            'emergency_number': (str(data.get('emergency_number')).strip() if data.get('emergency_number') else None),
            'status': status_norm,
            'manual_balance': parsed_manual
        }

        rows.append({'row': i, 'data': cleaned, 'errors': row_errors})

    return rows, errors

@app.route('/employees/import', methods=['GET', 'POST'])
@login_required
def import_employees():
    if not has_permission('can_edit_employee') and session.get('role') not in ('developer','admin_master'):
        flash('Import employees requires permission to edit employees', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file uploaded', 'danger')
            return redirect(url_for('import_employees'))
        # only accept xlsx
        if not f.filename.lower().endswith('.xlsx'):
            flash('Only .xlsx files accepted', 'danger')
            return redirect(url_for('import_employees'))

        fname = secure_filename(f.filename)
        uid = uuid.uuid4().hex
        dest = os.path.join(app.config['UPLOAD_FOLDER'], f'import_{uid}_{fname}')
        f.save(dest)

        rows, file_errors = parse_xlsx_file(dest)
        if file_errors:
            flash('File errors: ' + '; '.join(file_errors), 'danger')
            return redirect(url_for('import_employees'))

        to_create = []
        to_reject = []
        for r in rows:
            empid = r['data']['employee_id']
            if Employee.query.filter_by(employee_id=empid).first():
                r['errors'].append('Employee exists in database (duplicate). Use Edit or enable update mode.')
                to_reject.append(r)
            elif r['errors']:
                to_reject.append(r)
            else:
                to_create.append(r)

        preview_token = uuid.uuid4().hex
        preview_path = os.path.join(app.config['UPLOAD_FOLDER'], f'preview_{preview_token}.json')
        with open(preview_path, 'w', encoding='utf-8') as fh:
            json.dump({
                'source_filename': fname,
                'uploaded_path': dest,
                'rows': rows
            }, fh, default=str)

        return render_template('import_preview.html',
                               preview_token=preview_token,
                               to_create=to_create,
                               to_reject=to_reject,
                               total=len(rows),
                               created_count=len(to_create),
                               rejected_count=len(to_reject))

    return render_template('import_employees.html')

@app.route('/employees/import/confirm', methods=['POST'])
@login_required
def import_employees_confirm():
    if not has_permission('can_edit_employee') and session.get('role') not in ('developer','admin_master'):
        flash('Import confirm requires permission to edit employees', 'danger')
        return redirect(url_for('index'))

    token = request.form.get('preview_token')
    if not token:
        flash('Missing preview token', 'danger')
        return redirect(url_for('import_employees'))

    preview_path = os.path.join(app.config['UPLOAD_FOLDER'], f'preview_{token}.json')
    if not os.path.exists(preview_path):
        flash('Import preview data not found or expired', 'danger')
        return redirect(url_for('import_employees'))

    with open(preview_path, 'r', encoding='utf-8') as fh:
        payload = json.load(fh)

    rows = payload.get('rows', [])

    try:
        timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'pre_import_{timestamp}.db'
        backup_path = os.path.join(BACKUPS_FOLDER, backup_name)
        if os.path.exists(DB_PATH):
            shutil.copy2(DB_PATH, backup_path)
    except Exception as e:
        flash('Failed to create pre-import backup: ' + str(e), 'danger')
        return redirect(url_for('import_employees'))

    created = []
    failed = []
    try:
        for r in rows:
            if r.get('errors'):
                failed.append({'row': r['row'], 'errors': r['errors']})
                continue

            d = r['data']

            hire_val = d.get('hire_date')
            parsed_hire = None
            if hire_val in (None, ''):
                parsed_hire = None
            else:
                if isinstance(hire_val, str):
                    hire_val = hire_val.strip()
                    try:
                        parsed_hire = dt.strptime(hire_val, '%Y-%m-%d').date()
                    except Exception:
                        try:
                            parsed_hire = dt.strptime(hire_val, '%d-%m-%Y').date()
                        except Exception:
                            try:
                                parsed_hire = dt.fromisoformat(hire_val).date()
                            except Exception:
                                parsed_hire = None
                elif isinstance(hire_val, (dt,)):
                    parsed_hire = hire_val.date()
                else:
                    if hasattr(hire_val, 'year') and hasattr(hire_val, 'month'):
                        parsed_hire = hire_val
                    else:
                        parsed_hire = None

            accr_val = d.get('accrual_rate')
            parsed_accr = None
            try:
                parsed_accr = float(accr_val) if accr_val not in (None, '') else None
            except Exception:
                parsed_accr = None

            manual_val = d.get('manual_balance')
            parsed_manual = None
            if manual_val not in (None, ''):
                try:
                    parsed_manual = float(manual_val)
                except Exception:
                    parsed_manual = None

            status_val = d.get('status') or 'active'
            if isinstance(status_val, str):
                status_norm = status_val.strip().lower()
                if status_norm == '':
                    status_norm = 'active'
                if status_norm not in ('active', 'left'):
                    status_norm = 'active'
            else:
                status_norm = 'active'

            safe = {
                'employee_id': d.get('employee_id'),
                'first_name': d.get('first_name'),
                'middle_name': d.get('middle_name'),
                'last_name': d.get('last_name'),
                'hire_date': parsed_hire,
                'accrual_rate': parsed_accr,
                'status': status_norm,
                'plant_location': d.get('plant_location'),
                'department': d.get('department'),
                'designation': d.get('designation'),
                'contact_number': d.get('contact_number'),
                'emergency_number': d.get('emergency_number'),
                'manual_balance': parsed_manual
            }

            if Employee.query.filter_by(employee_id=safe['employee_id']).first():
                failed.append({'row': r['row'], 'errors': ['Already exists in DB']})
                continue

            # create employee
            emp = Employee(
                employee_id = safe['employee_id'],
                first_name = safe['first_name'],
                middle_name = safe.get('middle_name'),
                last_name = safe['last_name'],
                hire_date = safe['hire_date'],
                accrual_rate = safe['accrual_rate'],
                status = safe.get('status') or 'active',
                plant_location = safe.get('plant_location'),
                department = safe.get('department'),
                designation = safe.get('designation'),
                contact_number = safe.get('contact_number'),
                emergency_number = safe.get('emergency_number'),
            )
            db.session.add(emp)
            db.session.flush()  

            if safe.get('manual_balance') is not None:
                try:
                    try:
                        current_balance = compute_balance(emp)
                    except Exception:
                        current_balance = 0.0
                    delta = round(float(safe['manual_balance']) - float(current_balance), 2)
                    tr = Transaction(employee_id=emp.id, type='MANUAL_OVERRIDE', period=None,
                                     amount=delta, reference_id=None,
                                     note=f'Initial manual balance set via import by {session.get("username")}',
                                     created_by=session.get('user_id'))
                    db.session.add(tr)
                except Exception:
                    failed.append({'row': r['row'], 'errors': ['manual_balance audit failed']})

            created.append({'row': r['row'], 'employee_id': safe['employee_id']})

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash('Import failed and rolled back: ' + str(e), 'danger')
        return redirect(url_for('import_employees'))

    try:
        tr_summary = Transaction(employee_id=None, type='IMPORT', period=None,
                                 amount=0.0, reference_id=None,
                                 note=f'Imported {len(created)} employees from file {payload.get("source_filename")} by {session.get("username")}',
                                 created_by=session.get("user_id"))
        db.session.add(tr_summary)
        db.session.commit()
    except Exception:
        db.session.rollback()

    # write result CSV and offer download (simple)
    result_path = os.path.join(app.config['UPLOAD_FOLDER'], f'import_result_{token}.csv')
    with open(result_path, 'w', newline='', encoding='utf-8') as rf:
        w = csv.writer(rf)
        w.writerow(['row','employee_id','status','message'])
        for c in created:
            w.writerow([c['row'], c['employee_id'], 'created', 'OK'])
        for frow in failed:
            w.writerow([frow.get('row'), '', 'failed', ';'.join(frow.get('errors') or [])])

    flash(f'Import complete: created={len(created)}, failed={len(failed)}. Download results if needed.', 'success')
    return redirect(url_for('import_employees'))

# ------------------ Comp-Off routes ------------------

@app.route('/comp_off/new', methods=['GET', 'POST'])
@login_required
def comp_off_new():
    allowed_roles = ('developer', 'admin_master', 'admin_1', 'hr', 'hr_admin')
    if session.get('role') not in allowed_roles and not has_permission('can_edit_employee'):
        flash('Insufficient permission to create comp-off records.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        emp_code = request.form.get('employee_id','').strip()
        if not emp_code:
            flash('Employee ID required', 'danger')
            return redirect(url_for('comp_off_new'))

        emp = Employee.query.filter_by(employee_id=emp_code).first()
        if not emp:
            flash('Employee not found', 'danger')
            return redirect(url_for('comp_off_new'))

        if is_employee_left(emp, date.today()):
            left_str = getattr(emp, 'left_date', None)
            flash(f'Employee {emp_code} has left (on {left_str}) — cannot record comp-off.', 'warning')
            return redirect(url_for('comp_off_new'))

        earned_on_raw = request.form.get('earned_on','').strip()
        taken_on_raw = request.form.get('taken_on','').strip()
        try:
            earned_on = dt.strptime(earned_on_raw, '%Y-%m-%d').date()
        except Exception:
            flash('Invalid Earned On date. Use YYYY-MM-DD.', 'danger')
            return redirect(url_for('comp_off_new'))

        taken_on = None
        if taken_on_raw:
            try:
                taken_on = dt.strptime(taken_on_raw, '%Y-%m-%d').date()
            except Exception:
                flash('Invalid Taken On date. Use YYYY-MM-DD.', 'danger')
                return redirect(url_for('comp_off_new'))

        approved_by = request.form.get('approved_by','').strip() or None
        note = request.form.get('note','').strip() or None

        cor = CompOffRecord(
            employee_id=emp.id,
            emp_code=emp.employee_id,
            emp_name=' '.join(filter(None, [emp.first_name, getattr(emp,'middle_name',None) or '', emp.last_name])),
            department=getattr(emp, 'department', None),
            earned_on=earned_on,
            taken_on=taken_on,
            approved_by=approved_by,
            note=note,
            created_by=session.get('user_id')
        )
        db.session.add(cor)
        db.session.commit()

        flash(
            f"Comp-Off recorded successfully for {emp.employee_id}", "success"
        )
        return redirect(url_for('comp_off_new'))
    return render_template('comp_off_form.html')

@app.route('/comp_off/list')
@login_required
def comp_off_list():
    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()
    q_year = request.args.get('year','').strip()
    q_month = request.args.get('month','').strip()

    q = CompOffRecord.query.order_by(CompOffRecord.earned_on.desc())
    if q_emp:
        q = q.filter(CompOffRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(CompOffRecord.department == q_dept)
    if q_year:
        try:
            y = int(q_year)
            q = q.filter(db.extract('year', CompOffRecord.earned_on) == y)
        except Exception:
            pass
    if q_month:
        try:
            m = int(q_month)
            q = q.filter(db.extract('month', CompOffRecord.earned_on) == m)
        except Exception:
            pass

    page = int(request.args.get('page', 1))
    per_page = 50
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    compoffs = pagination.items

    return render_template('comp_off_list.html',
                           compoffs=compoffs,
                           pagination=pagination,
                           q_emp=q_emp, q_dept=q_dept, q_year=q_year, q_month=q_month)


@app.route('/comp_off/export')
@login_required
def comp_off_export():
    fmt = request.args.get('format','csv')

    today = date.today()
    first_day = date(today.year, today.month, 1)
    if today.month == 12:
        last_day = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(today.year, today.month + 1, 1) - timedelta(days=1)

    q = CompOffRecord.query.filter(
        CompOffRecord.earned_on >= first_day,
        CompOffRecord.earned_on <= last_day
    ).order_by(CompOffRecord.earned_on.asc())

    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()

    if q_emp:
        q = q.filter(CompOffRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(CompOffRecord.department == q_dept)

    rows = []
    for co in q.all():
        rows.append({
            'EMPLOYEE_ID': co.emp_code,
            'EMPLOYEE_NAME': co.emp_name,
            'DEPARTMENT': co.department or '',
            'EARNED_ON': co.earned_on.strftime('%d/%m/%Y') if co.earned_on else '',
            'TAKEN_ON': co.taken_on.strftime('%d/%m/%Y') if co.taken_on else '',
            'APPROVED_BY': co.approved_by or '',
            'NOTE': co.note or ''
        })

    if fmt == 'xlsx':
        df = pd.DataFrame(rows)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='CompOffs')
        bio.seek(0)
        return send_file(bio,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name=f"comp_offs_{dt.utcnow().strftime('%Y%m%d')}.xlsx",
                         as_attachment=True)
    else:
        si = StringIO()
        if rows:
            writer = csv.DictWriter(si, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
        output = si.getvalue().encode('utf-8')
        return send_file(BytesIO(output),
                         mimetype='text/csv',
                         as_attachment=True,
                         download_name=f"comp_offs_{dt.utcnow().strftime('%Y%m%d')}.csv")
# ------------------ end Comp-Off routes ------------------

# ------------------ Early/Late routes ------------------
@app.route('/early_late/new', methods=['GET', 'POST'])
@login_required
def early_late_record():
    # permission: adjust role list as needed
    allowed_roles = ('developer', 'admin_master', 'admin_1', 'hr', 'hr_admin')
    if session.get('role') not in allowed_roles and not has_permission('can_edit_employee'):
        flash('Insufficient permission to create early/late records.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        emp_code = request.form.get('employee_id','').strip()
        if not emp_code:
            flash('Employee ID required', 'danger')
            return redirect(url_for('early_late_record'))

        emp = Employee.query.filter_by(employee_id=emp_code).first()
        if not emp:
            flash('Employee not found', 'danger')
            return redirect(url_for('early_late_record'))

        # block if left
        if is_employee_left(emp, date.today()):
            left_str = getattr(emp, 'left_date', None)
            flash(f'Employee {emp_code} has left (on {left_str}) — cannot record.', 'warning')
            return redirect(url_for('early_late_record'))

        late_raw = request.form.get('late_datetime','').strip()
        early_raw = request.form.get('early_datetime','').strip()

        if not late_raw and not early_raw:
            flash('Please enter Late Coming OR Early Going time.', 'danger')
            return redirect(url_for('early_late_record'))

        late_dt = None
        early_dt = None
        try:
            if late_raw:
                late_dt = dt.strptime(late_raw, '%Y-%m-%dT%H:%M')
        except Exception:
            flash('Invalid Late coming datetime (use the picker).', 'danger')
            return redirect(url_for('early_late_record'))
        try:
            if early_raw:
                early_dt = dt.strptime(early_raw, '%Y-%m-%dT%H:%M')
        except Exception:
            flash('Invalid Early going datetime (use the picker).', 'danger')
            return redirect(url_for('early_late_record'))

        approver = request.form.get('approver','').strip() or None
        note = request.form.get('note','').strip() or None

        el = EarlyLateRecord(
            employee_id = emp.id,
            emp_code = emp.employee_id,
            emp_name = ' '.join(filter(None, [emp.first_name, getattr(emp,'middle_name',None) or '', emp.last_name])),
            department = getattr(emp, 'department', None),
            late_datetime = late_dt,
            early_datetime = early_dt,
            approved_by = approver,   
            note = note,
            created_by = session.get('user_id')
        )

        db.session.add(el)
        db.session.commit()

        flash(f'Early/Late record saved for {emp.employee_id}', 'success')
        return redirect(url_for('early_late_record'))

    return render_template('early_late.html')


@app.route('/early_late/list')
@login_required
def early_late_list():
    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()
    q_year = request.args.get('year','').strip()
    q_month = request.args.get('month','').strip()

    q = EarlyLateRecord.query.order_by(EarlyLateRecord.late_datetime.desc().nullslast(), EarlyLateRecord.early_datetime.desc().nullslast())

    if q_emp:
        q = q.filter(EarlyLateRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(EarlyLateRecord.department == q_dept)
    if q_year:
        try:
            y = int(q_year)
            q = q.filter(
                or_(
                    db.extract('year', EarlyLateRecord.late_datetime) == y,
                    db.extract('year', EarlyLateRecord.early_datetime) == y
                )
            )
        except Exception:
            pass

    if q_month:
        try:
            m = int(q_month)
            q = q.filter(
                or_(
                    db.extract('month', EarlyLateRecord.late_datetime) == m,
                    db.extract('month', EarlyLateRecord.early_datetime) == m
                )
            )
        except Exception:
            pass

    page = int(request.args.get('page', 1))
    per_page = 50
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    return render_template('early_late_list.html',
                           items=items,
                           pagination=pagination,
                           q_emp=q_emp, q_dept=q_dept, q_year=q_year, q_month=q_month)


@app.route('/early_late/export')
@login_required
def early_late_export():
    fmt = request.args.get('format','csv')
    today = date.today()
    current_year = today.year
    current_month = today.month
    q = EarlyLateRecord.query.filter(
        or_(
            and_(
                EarlyLateRecord.late_datetime.isnot(None),
                db.extract('year', EarlyLateRecord.late_datetime) == current_year,
                db.extract('month', EarlyLateRecord.late_datetime) == current_month
            ),
            and_(
                EarlyLateRecord.early_datetime.isnot(None),
                db.extract('year', EarlyLateRecord.early_datetime) == current_year,
                db.extract('month', EarlyLateRecord.early_datetime) == current_month
            )
        )
    ).order_by(
        EarlyLateRecord.late_datetime.asc().nullslast(),
        EarlyLateRecord.early_datetime.asc().nullslast()
    )

    q_emp = request.args.get('employee_id','').strip()
    q_dept = request.args.get('department','').strip()
    q_year = request.args.get('year','').strip()
    q_month = request.args.get('month','').strip()

    if q_emp:
        q = q.filter(EarlyLateRecord.emp_code.ilike(f'%{q_emp}%'))
    if q_dept:
        q = q.filter(EarlyLateRecord.department == q_dept)
    if q_year:
        try:
            y = int(q_year)
            q = q.filter(
                or_(
                    db.extract('year', EarlyLateRecord.late_datetime) == y,
                    db.extract('year', EarlyLateRecord.early_datetime) == y
                )
                
            )
        except Exception:
            pass

    if q_month:
        try:
            m = int(q_month)
            q = q.filter(
                or_(
                    db.extract('month', EarlyLateRecord.late_datetime) == m,
                    db.extract('month', EarlyLateRecord.early_datetime) == m
                )
            )
        except Exception:
            pass

    rows = []
    for r in q.all():
        rows.append({
            'EMPLOYEE_ID': r.emp_code,
            'EMPLOYEE_NAME': r.emp_name,
            'DEPARTMENT': r.department or '',
            'DATE': (
                r.late_datetime.strftime('%d/%m/%Y')
                if r.late_datetime else
                r.early_datetime.strftime('%d/%m/%Y')
                if r.early_datetime else ''
            ),
            'LATE_TIME': r.late_datetime.strftime('%H:%M') if r.late_datetime else '',
            'EARLY_TIME': r.early_datetime.strftime('%H:%M') if r.early_datetime else '',
            'NOTE': r.note or ''
        })


    if fmt == 'xlsx':
        df = pd.DataFrame(rows)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='EarlyLate')
        bio.seek(0)
        return send_file(bio,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name=f"early_late_{dt.utcnow().strftime('%Y%m%d')}.xlsx",
                         as_attachment=True)
    else:
        si = StringIO()
        if rows:
            writer = csv.DictWriter(si, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for r in rows:
                writer.writerow(r)

        output = si.getvalue().encode('utf-8')
        return send_file(
            BytesIO(output),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"early_late_{dt.utcnow().strftime('%Y%m%d')}.csv"
        )

# ------------------ end Early/Late routes ------------------

# ------------------ Start OD Outdoor Duty routes ------------------

@app.route('/outdoor/new', methods=['GET', 'POST'])
@login_required
def outdoor_new():
    allowed_roles = ('developer', 'admin_master', 'admin_1')
    if session.get('role') not in allowed_roles:
        flash('Permission denied.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        # ---------- EMPLOYEE ----------
        emp_code = request.form.get('employee_id', '').strip()
        emp = Employee.query.filter_by(employee_id=emp_code).first()

        if not emp:
            flash('Employee not found', 'danger')
            return redirect(url_for('outdoor_new'))

        if is_employee_left(emp):
            flash('Employee has left.', 'danger')
            return redirect(url_for('outdoor_new'))

        # ---------- OD TYPE ----------
        od_type = request.form.get('od_type')
        is_full_day = (od_type == 'full')

        # ---------- FULL DAY ----------
        if is_full_day:
            od_date_raw = request.form.get('od_date')

            if not od_date_raw:
                flash('OD date is required', 'danger')
                return redirect(url_for('outdoor_new'))

            try:
                od_date = dt.strptime(od_date_raw, '%Y-%m-%d').date()
            except Exception:
                flash('Invalid OD date', 'danger')
                return redirect(url_for('outdoor_new'))

            time_from = None
            time_to = None

        # ---------- HALF DAY ----------
        else:
            from_raw = request.form.get('from_datetime')
            to_raw = request.form.get('to_datetime')

            if not from_raw or not to_raw:
                flash('From and To time required for Half Day OD', 'danger')
                return redirect(url_for('outdoor_new'))

            try:
                time_from = dt.strptime(from_raw, '%Y-%m-%dT%H:%M')
                time_to = dt.strptime(to_raw, '%Y-%m-%dT%H:%M')
            except Exception:
                flash('Invalid date/time format', 'danger')
                return redirect(url_for('outdoor_new'))

            if time_to <= time_from:
                flash('To time must be after From time', 'danger')
                return redirect(url_for('outdoor_new'))

            od_date = time_from.date()

        # ---------- SAVE ----------
        od = OutdoorDuty(
            employee_id=emp.id,
            emp_code=emp.employee_id,
            emp_name=f"{emp.first_name} {emp.last_name}",
            department=emp.department,
            designation=emp.designation,
            od_date=od_date,
            is_full_day=is_full_day,
            time_from=time_from,
            time_to=time_to,
            reason=request.form.get('reason'),
            approved_by=request.form.get('approved_by','').strip() or None,   # ✅ FIXED
            note=request.form.get('note'),
            created_by=session.get('user_id')
        )

        db.session.add(od)
        db.session.commit()

        flash('Outdoor duty recorded', 'success')
        return redirect(url_for('history_outdoor'))

    return render_template('outdoor_duty_form.html')


# ------------------ End OD Outdoor Duty routes ------------------

@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html", hide_topbar=True)


# ---------- Run server ----------
if __name__ == "__main__":
    with app.app_context():
        try:
            User.query.update({User.session_token: None})
            db.session.commit()
        except Exception:
            pass
    app.run(host="127.0.0.1", port=5000, debug=True)
